from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.services.processor import FACTORY_COLUMNS, GAP_COLUMN, REQUIRED_COLUMNS, SUMMARY_COLUMNS, ProcessingError, process_workbook


ALL_SOURCE_COLUMNS = REQUIRED_COLUMNS + ["ExpectedDiaWt", "Total Value"]


def workbook_bytes(headers=ALL_SOURCE_COLUMNS, rows=None):
    book = Workbook()
    sheet = book.active
    sheet.append(headers)
    for row in rows or []:
        sheet.append(row)
    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()


def sample_row(delivery=date(2026, 8, 15), factory="OB1", bag_qty=10, weight=2.5, value=100):
    return [date(2026, 8, 1), "C1", "S1", bag_qty, "001245", delivery, factory, "VS", weight, value]


def test_missing_required_column():
    headers = [c for c in ALL_SOURCE_COLUMNS if c != "ExpDiaDlvDate"]
    with pytest.raises(ProcessingError, match="ExpDiaDlvDate"):
        process_workbook(workbook_bytes(headers, [[1] * len(headers)]))


@pytest.mark.parametrize("delivery,expected", [
    (date(2026, 8, 15), 6),
    (date(2026, 8, 18), 3),
    (date(2026, 8, 21), None),
    (date(2026, 8, 25), None),
    (None, None),
])
def test_days_gap_only_for_overdue_dates(delivery, expected):
    result = process_workbook(workbook_bytes(rows=[sample_row(delivery)]), today=date(2026, 8, 21))
    assert result.records[0][GAP_COLUMN] == expected


def test_excel_output_has_exact_tabs_and_columns():
    result = process_workbook(workbook_bytes(rows=[sample_row()]), today=date(2026, 8, 21))
    book = load_workbook(BytesIO(result.content), data_only=False)
    assert book.sheetnames == ["Processed Data", "Detail summary", "Factory Summary"]
    assert [c.value for c in book["Processed Data"][1]] == REQUIRED_COLUMNS
    assert [c.value for c in book["Detail summary"][1]] == SUMMARY_COLUMNS
    assert [c.value for c in book["Factory Summary"][1]] == FACTORY_COLUMNS
    assert book["Detail summary"]["F2"].value == '=IF(AND(ISNUMBER(A2),TODAY()>A2),TODAY()-A2,"")'


def test_factory_summary_aggregates_factory_totals():
    rows = [sample_row(factory="Factory A", bag_qty=10, weight=2.5, value=100), sample_row(factory="Factory A", bag_qty=5, weight=1.25, value=40)]
    result = process_workbook(workbook_bytes(rows=rows), today=date(2026, 8, 21))
    assert result.factory_summary == [{"Factory Name": "Factory A", "Total No. of Bag Qty": 15, "Total Sum of Expected Dia Wt": 3.75, "Total Sum Value": 140}]


def test_invalid_delivery_date_warns_and_stays_blank():
    result = process_workbook(workbook_bytes(rows=[sample_row("not-a-date")]), today=date(2026, 8, 21))
    assert result.records[0][GAP_COLUMN] is None
    assert "invalid ExpDiaDlvDate" in result.warnings[0]


def test_formula_like_identifier_is_written_as_text():
    row = sample_row()
    row[1] = '=HYPERLINK("bad")'
    result = process_workbook(workbook_bytes(rows=[row]), today=date(2026, 8, 21))
    book = load_workbook(BytesIO(result.content), data_only=False)
    assert book["Processed Data"]["B2"].data_type == "s"
    assert book["Processed Data"]["B2"].value.startswith("'=")
