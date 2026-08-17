from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REQUIRED_COLUMNS = [
    "OrderDate",
    "ClientCode",
    "StyleCode",
    "BagQty",
    "PoNo",
    "ExpDiaDlvDate",
    "Order Book",
    "Expected Dia Qly",
]
SUMMARY_COLUMNS = ["Date", "Order Book", "Client name", "Bag Qty", "PO No.", "Gap Days"]
FACTORY_COLUMNS = ["Factory Name", "Total No. of Bag Qty", "Total Sum of Expected Dia Wt", "Total Sum Value"]
FACTORY_SOURCE_COLUMNS = ["ExpectedDiaWt", "Total Value"]
GAP_COLUMN = "Gap Days"

# Deliberately conservative: aliases can be added here without changing processing code.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {name: (name,) for name in REQUIRED_COLUMNS}
IDENTIFIER_COLUMNS = {"ClientCode", "StyleCode", "PoNo", "Order Book", "Expected Dia Qly"}
DATE_COLUMNS = {"OrderDate", "ExpDiaDlvDate"}


class ProcessingError(ValueError):
    pass


@dataclass
class ProcessingResult:
    content: bytes
    records: list[dict[str, Any]]
    factory_summary: list[dict[str, Any]]
    warnings: list[str]
    total_records: int


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _column_map(headers: list[Any]) -> dict[str, int]:
    normalized = {_normalize_header(value): index for index, value in enumerate(headers, start=1)}
    found: dict[str, int] = {}
    missing: list[str] = []
    for required, aliases in COLUMN_ALIASES.items():
        match = next((normalized[_normalize_header(alias)] for alias in aliases if _normalize_header(alias) in normalized), None)
        if match is None:
            missing.append(required)
        else:
            found[required] = match
    for required in FACTORY_SOURCE_COLUMNS:
        match = normalized.get(_normalize_header(required))
        if match is None:
            missing.append(required)
        else:
            found[required] = match
    if missing:
        label = "column" if len(missing) == 1 else "columns"
        raise ProcessingError(f"The following required {label} {'is' if len(missing) == 1 else 'are'} missing: {', '.join(missing)}")
    return found


def _date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _identifier_value(value: Any, number_format: str) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int) or (isinstance(value, float) and value.is_integer()):
        numeric = int(value)
        if number_format and re.fullmatch(r"0+", number_format):
            return f"{numeric:0{len(number_format)}d}"
        return str(numeric)
    return str(value).strip()


def _bag_qty(value: Any) -> int | float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    try:
        number = float(str(value).replace(",", "").strip())
        return int(number) if number.is_integer() else number
    except (TypeError, ValueError):
        return None


def _display(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.strftime("%d-%b-%Y")
    return value


def _safe_excel_value(value: Any) -> Any:
    """Keep uploaded text as text when it resembles an Excel formula."""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _style_sheet(ws, columns: int, rows: int) -> None:
    header_fill = PatternFill("solid", fgColor="17324D")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(columns)}{max(rows, 1)}"
    ws.row_dimensions[1].height = 24
    widths = [18, 24, 24, 13, 20, 16, 24, 24]
    for index in range(1, columns + 1):
        ws.column_dimensions[get_column_letter(index)].width = widths[index - 1]
    if rows > 1:
        table = Table(displayName=f"{ws.title.replace(' ', '')}Table", ref=f"A1:{get_column_letter(columns)}{rows}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws.add_table(table)


def process_workbook(content: bytes, today: date | None = None, max_rows: int = 100000) -> ProcessingResult:
    today = today or date.today()
    try:
        source = load_workbook(BytesIO(content), read_only=False, data_only=False)
    except Exception as exc:
        raise ProcessingError("The uploaded file is not a valid Excel workbook.") from exc
    if not source.worksheets:
        raise ProcessingError("The Excel workbook does not contain a worksheet.")

    ws = source.worksheets[0]
    if ws.max_row < 1:
        raise ProcessingError("The first worksheet is empty.")
    headers = [cell.value for cell in ws[1]]
    mapping = _column_map(headers)
    records: list[dict[str, Any]] = []
    warnings: list[str] = []

    for row_number in range(2, ws.max_row + 1):
        raw = {name: ws.cell(row_number, column).value for name, column in mapping.items()}
        if all(value in (None, "") for value in raw.values()):
            continue
        if len(records) >= max_rows:
            raise ProcessingError(f"The file exceeds the maximum of {max_rows:,} data rows.")
        item: dict[str, Any] = {}
        for name in REQUIRED_COLUMNS:
            cell = ws.cell(row_number, mapping[name])
            if name in DATE_COLUMNS:
                parsed = _date_value(cell.value)
                item[name] = parsed
                if cell.value not in (None, "") and parsed is None:
                    warnings.append(f"Row {row_number}: invalid {name}; the value was left blank.")
            elif name == "BagQty":
                item[name] = _bag_qty(cell.value)
                if cell.value not in (None, "") and item[name] is None:
                    warnings.append(f"Row {row_number}: invalid BagQty; the value was left blank.")
            else:
                item[name] = _identifier_value(cell.value, cell.number_format)
        delivery = item["ExpDiaDlvDate"]
        gap = (today - delivery).days if delivery else None
        item[GAP_COLUMN] = gap if gap is not None and gap > 0 else None
        item["ExpectedDiaWt"] = _bag_qty(ws.cell(row_number, mapping["ExpectedDiaWt"]).value) or 0
        item["Total Value"] = _bag_qty(ws.cell(row_number, mapping["Total Value"]).value) or 0
        records.append(item)

    if not records:
        raise ProcessingError("The first worksheet has no data rows to process.")

    output = Workbook()
    processed = output.active
    processed.title = "Processed Data"
    summary = output.create_sheet("Detail summary")
    factory_sheet = output.create_sheet("Factory Summary")
    processed.append(REQUIRED_COLUMNS)
    summary.append(SUMMARY_COLUMNS)
    factory_sheet.append(FACTORY_COLUMNS)
    for excel_row, item in enumerate(records, start=2):
        processed.append([_safe_excel_value(item[name]) for name in REQUIRED_COLUMNS])
        summary.append([
            item["ExpDiaDlvDate"],
            _safe_excel_value(item["Order Book"]),
            _safe_excel_value(item["ClientCode"]),
            item["BagQty"],
            _safe_excel_value(item["PoNo"]),
            None,
        ])
        summary.cell(excel_row, 6, f'=IF(AND(ISNUMBER(A{excel_row}),TODAY()>A{excel_row}),TODAY()-A{excel_row},"")')
    factory_totals: dict[str, dict[str, float]] = {}
    for item in records:
        factory = item["Order Book"] or "(Blank)"
        totals = factory_totals.setdefault(factory, {"bags": 0, "weight": 0, "value": 0})
        totals["bags"] += item["BagQty"] or 0
        totals["weight"] += item["ExpectedDiaWt"]
        totals["value"] += item["Total Value"]
    factory_summary = []
    for factory in sorted(factory_totals, key=str.casefold):
        totals = factory_totals[factory]
        row = {"Factory Name": factory, "Total No. of Bag Qty": totals["bags"], "Total Sum of Expected Dia Wt": totals["weight"], "Total Sum Value": totals["value"]}
        factory_summary.append(row)
        factory_sheet.append(list(row.values()))
    for target in (processed, summary, factory_sheet):
        date_columns = (1, 6) if target.title == "Processed Data" else (1,)
        for col in date_columns:
            for row in range(2, target.max_row + 1):
                target.cell(row, col).number_format = "dd-mmm-yyyy"
        for row in range(2, target.max_row + 1):
            target.cell(row, 4).number_format = "#,##0.###"
        _style_sheet(target, target.max_column, target.max_row)
    for row in range(2, factory_sheet.max_row + 1):
        factory_sheet.cell(row, 2).number_format = "#,##0"
        factory_sheet.cell(row, 3).number_format = "#,##0"
        factory_sheet.cell(row, 4).number_format = "#,##0"

    buffer = BytesIO()
    output.save(buffer)
    preview = []
    for item in records:
        view = dict(item)
        view.update({
            "Date": item["ExpDiaDlvDate"],
            "Client name": item["ClientCode"],
            "Bag Qty": item["BagQty"],
            "PO No.": item["PoNo"],
        })
        preview.append({key: _display(value) for key, value in view.items()})
    return ProcessingResult(buffer.getvalue(), preview, factory_summary, warnings, len(records))
